"""
Dry-run preview of the EINV & WSPP (SST) xlsx sheet import.
READS ONLY — no writes to Supabase.

Usage:
  python scripts/einv-dry-run.py "CRM 220426.xlsx"

Output:
  - Sheet detection summary
  - Match/orphan counts between CRM sheet, EINV sheet, and live clinics table
  - Per-field change preview (first 20 clinics)
  - Summary totals per field
"""

import os
import sys
import re
from pathlib import Path
from collections import defaultdict
from datetime import datetime, date

import openpyxl
import requests
from dotenv import load_dotenv

# Load env
root = Path(__file__).resolve().parent.parent
load_dotenv(root / ".env.local")

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL")
SERVICE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

if not SUPABASE_URL or not SERVICE_KEY:
    sys.exit("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "CRM 220426.xlsx"
if not Path(XLSX_PATH).exists():
    sys.exit(f"File not found: {XLSX_PATH}")

# --- Helpers (mirrors the TS route's fixDate / toStr / boolean parser) ---

def to_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Filter Excel errors
    if s in ("#REF!", "#N/A", "#VALUE!", "#NAME?", "#DIV/0!"):
        return None
    return s

def to_bool(val):
    """Returns True, False, or None. None = unknown (blank / error)."""
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("true", "1", "yes", "y", "✓"):
        return True
    if s in ("false", "0", "no", "n", ""):
        return False
    if s in ("#ref!", "#n/a", "#value!"):
        return None
    return None

def fix_date(val):
    """Returns YYYY-MM-DD string or None."""
    if val is None or val == "":
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    # DD/MM/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return None

# --- Read xlsx ---

print(f"Reading {XLSX_PATH}...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

# Find sheets (handle trailing-space bug on EINV sheet)
crm_sheet_name = next((n for n in wb.sheetnames if n.strip() == "CRM"), None)
einv_sheet_name = next((n for n in wb.sheetnames if n.strip() == "EINV & WSPP (SST)"), None)

if not crm_sheet_name:
    sys.exit("CRM sheet not found")
if not einv_sheet_name:
    sys.exit("EINV & WSPP (SST) sheet not found")

print(f"  CRM sheet: '{crm_sheet_name}'")
print(f"  EINV sheet: '{einv_sheet_name}'  (trailing-space: {einv_sheet_name != einv_sheet_name.strip()})")
print()

# --- Read CRM sheet ---
ws = wb[crm_sheet_name]
crm_headers = [to_str(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
crm_acct_idx = crm_headers.index("ACCT NO")
crm_codes = set()
for r in ws.iter_rows(min_row=2, values_only=True):
    code = to_str(r[crm_acct_idx]) if crm_acct_idx < len(r) else None
    if code:
        crm_codes.add(code)

# --- Read EINV sheet (headers on row 2, data from row 3) ---
ws = wb[einv_sheet_name]
einv_headers_raw = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))[:23]
# Normalize headers: collapse \n and whitespace (matches TS route's normalizeHeaders)
einv_headers = [re.sub(r"\s+", " ", str(h).strip()) if h else None for h in einv_headers_raw]

# Column index map (by normalized header)
def col_idx(name):
    try:
        return einv_headers.index(name)
    except ValueError:
        print(f"  WARNING: EINV header not found: {name!r}")
        return None

EINV_MAP = {
    # xlsx normalized header : (db column, parser)
    "ACC NO": ("clinic_code", to_str),
    "SST Registered no": ("sst_registration_no", to_str),
    "Tarikh Kuatkuasa Pendaftaran": ("sst_start_date", to_str),  # existing col is text; keep as str
    "Tempoh bercukai (1 or 2mnth)": ("sst_frequency", to_str),
    "Tempoh Bercukai Beikutnya (1 or 2mnth)": ("sst_period_next", to_str),
    "E-INV V1 (RM699-setup)": ("einv_v1_signed", to_bool),
    "E-INV V2 (RM500-Yearly Hosting)": ("einv_v2_signed", to_bool),
    "WHATSAPP SETUP (RM500-Yearly Hosting)": ("has_whatsapp", to_bool),
    "STATUS INSTALLATION": ("einv_install_status", to_str),
    "Username PSW: Medexone@603 / Medex@603": ("einv_portal_credentials", to_str),
    "INSTALL DATE (E-INV V2)": ("einv_install_date", fix_date),
    "Set up fee RM699": ("einv_setup_fee_status", to_str),
    "Hosting fee status RM500": ("einv_hosting_fee_status", to_str),
    "Payment Date (only Hosting)": ("einv_payment_date", fix_date),
}

# Verify all headers found
missing = []
for hdr in EINV_MAP:
    if hdr not in einv_headers:
        missing.append(hdr)
if missing:
    print("Available EINV headers (normalized):")
    for i, h in enumerate(einv_headers):
        print(f"  [{i}] {h!r}")
    sys.exit(f"\nMissing headers: {missing}")

# Build EINV merge map: clinic_code -> {db_col: value}
einv_merge = {}
einv_rows_seen = 0
orphan_codes = []
ref_errors = 0

for r in ws.iter_rows(min_row=3, values_only=True):
    vals = list(r)[:23]
    if all(v in (None, "") for v in vals):
        continue
    einv_rows_seen += 1

    payload = {}
    for hdr, (db_col, parser) in EINV_MAP.items():
        idx = einv_headers.index(hdr)
        raw = vals[idx] if idx < len(vals) else None
        if raw == "#REF!":
            ref_errors += 1
            continue
        payload[db_col] = parser(raw)

    code = payload.get("clinic_code")
    if not code:
        continue

    # Derive has_e_invoice (any of V1/V2 true), has_sst (SST reg present)
    v1 = payload.get("einv_v1_signed")
    v2 = payload.get("einv_v2_signed")
    if v1 is True or v2 is True:
        payload["has_e_invoice"] = True
    elif v1 is False and v2 is False:
        payload["has_e_invoice"] = False
    # else leave None (unknown)

    payload["has_sst"] = payload.get("sst_registration_no") is not None

    einv_merge[code] = payload

# --- Fetch current clinics from Supabase ---
print(f"Fetching current clinics from Supabase...")
headers_api = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
}

# All fields we care about
fields = [
    "clinic_code",
    "clinic_name",
    "has_e_invoice", "has_sst", "has_whatsapp",
    "sst_registration_no", "sst_start_date", "sst_frequency", "sst_period_next",
    "einv_v1_signed", "einv_v2_signed",
    "einv_setup_fee_status", "einv_hosting_fee_status",
    "einv_payment_date", "einv_install_date",
    "einv_portal_credentials", "einv_install_status",
]

# Paginate (Supabase caps at 1000 per request)
current = {}
offset = 0
page_size = 1000
while True:
    url = f"{SUPABASE_URL}/rest/v1/clinics?select={','.join(fields)}&order=clinic_code"
    page_headers = {**headers_api, "Range": f"{offset}-{offset + page_size - 1}"}
    resp = requests.get(url, headers=page_headers)
    resp.raise_for_status()
    batch = resp.json()
    for c in batch:
        current[c["clinic_code"]] = c
    if len(batch) < page_size:
        break
    offset += page_size

print(f"  Supabase: {len(current)} clinics")
print(f"  CRM sheet: {len(crm_codes)} clinics")
print(f"  EINV sheet: {einv_rows_seen} data rows ({len(einv_merge)} unique codes)")
print(f"  #REF! errors: {ref_errors}")
print()

# --- Match / orphan analysis ---
matched = []
orphan = []
for code, payload in einv_merge.items():
    if code in current:
        matched.append(code)
    else:
        orphan.append(code)

print(f"=== Match analysis ===")
print(f"  EINV rows that match live clinic:     {len(matched)}")
print(f"  EINV rows with NO matching clinic:    {len(orphan)}")
if orphan:
    print(f"    Orphan codes (first 20): {orphan[:20]}")
print()

# Clinics in CRM sheet but not in Supabase (will be inserted)
new_to_supabase = crm_codes - set(current.keys())
# Clinics in Supabase but not in CRM sheet (will be stale-deleted by existing route logic)
dropping_out = set(current.keys()) - crm_codes
print(f"=== CRM sheet vs Supabase ===")
print(f"  New clinics (will be inserted):       {len(new_to_supabase)}")
print(f"  Clinics dropping out (will be deleted by stale cleanup): {len(dropping_out)}")
if dropping_out:
    print(f"    First 20: {sorted(dropping_out)[:20]}")
print()

# --- Per-field change count ---
field_changes = defaultdict(int)
changed_clinics = []  # list of (code, name, [(field, old, new)])

for code in matched:
    new_payload = einv_merge[code]
    cur = current[code]
    changes = []
    for fld, new_val in new_payload.items():
        if fld == "clinic_code":
            continue
        old_val = cur.get(fld)
        # Normalize for comparison
        if old_val != new_val:
            # Skip "None to None" (shouldn't happen but defensive)
            if old_val is None and new_val is None:
                continue
            changes.append((fld, old_val, new_val))
            field_changes[fld] += 1
    if changes:
        changed_clinics.append((code, cur.get("clinic_name", "?"), changes))

print(f"=== Field-level change totals (across {len(matched)} matched clinics) ===")
for fld in sorted(field_changes, key=lambda f: -field_changes[f]):
    print(f"  {fld:<30} {field_changes[fld]:>5}")
print()

print(f"=== Clinics that will change: {len(changed_clinics)} / {len(matched)} ===")
print(f"Clinics with no changes (already in sync): {len(matched) - len(changed_clinics)}")
print()

print(f"=== First 20 clinic change previews ===")
for code, name, changes in changed_clinics[:20]:
    print(f"\n  {code}  {name}")
    for fld, old, new in changes:
        print(f"    {fld:<26} {old!r:<30} -> {new!r}")

# Save full report
report_path = Path(XLSX_PATH).with_suffix(".dry-run.txt")
with open(report_path, "w", encoding="utf-8") as f:
    f.write(f"Dry-run report for {XLSX_PATH}\n")
    f.write(f"Generated: {datetime.now().isoformat()}\n\n")
    f.write(f"Supabase:   {len(current)} clinics\n")
    f.write(f"CRM sheet:  {len(crm_codes)} clinics\n")
    f.write(f"EINV sheet: {len(einv_merge)} unique codes (#REF! errors: {ref_errors})\n\n")
    f.write(f"Matched:    {len(matched)}\n")
    f.write(f"Orphans:    {len(orphan)}\n")
    f.write(f"New:        {len(new_to_supabase)}\n")
    f.write(f"Dropping:   {len(dropping_out)}\n\n")
    f.write(f"Will change: {len(changed_clinics)} clinics\n\n")
    f.write(f"Field-level change totals:\n")
    for fld, n in sorted(field_changes.items(), key=lambda x: -x[1]):
        f.write(f"  {fld:<30} {n:>5}\n")
    f.write(f"\n=== All changes ===\n")
    for code, name, changes in changed_clinics:
        f.write(f"\n{code}  {name}\n")
        for fld, old, new in changes:
            f.write(f"  {fld:<26} {old!r:<30} -> {new!r}\n")
    if orphan:
        f.write(f"\n=== Orphan EINV codes (not in Supabase) ===\n")
        for c in orphan:
            f.write(f"  {c}\n")
    if dropping_out:
        f.write(f"\n=== Clinics dropping out (will be stale-deleted) ===\n")
        for c in sorted(dropping_out):
            f.write(f"  {c}\n")

print(f"\nFull report saved to: {report_path}")
